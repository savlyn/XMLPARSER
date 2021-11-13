import numpy as np
import pandas as pd
import xml.etree.cElementTree as et
import os
import openpyxl
from openpyxl.utils.dataframe import dataframe_to_rows
import glob


# writing to excel
def savetoexcel(df, filename, sheet):
    #check to see if the file already exists, if yes check further if teh sheet already exists in the file
    if os.path.exists(filename):
        workbook = openpyxl.load_workbook(filename)
        if sheet in workbook.sheetnames:
            active_sheet = workbook[sheet]
            for row in dataframe_to_rows(df, header=False, index=False):
                active_sheet.append(row)

     #create new sheet 
        else:

            workbook.create_sheet(sheet)
            active_sheet = workbook[sheet]

            for row in dataframe_to_rows(df, index=False):
                active_sheet.append(row)

        workbook.save(filename)
        
    
    else:
        with pd.ExcelWriter(filename, engine='openpyxl', mode='w') as writer:

            # write dataframe to excel
            df.to_excel(writer, sheet_name=sheet, index=False)
            # save the excel
            writer.save()
          

# Function to Parse through the tree node to get data objects
def parser(name):
    os.chdir(name)
    for file in list(glob.glob('*.xml',recursive=True)):

        print('Executing file: {}'.format(file))
        datalist = ['./DataExtract900jer/Messages', './DataExtract900jer/Rules', './Nuula/Errors']
        tree = et.parse(file)
        root = tree.getroot()
        for i in datalist:
            (prefix, sep, suffix) = i.rpartition('/')
            for item in root.findall(i):
                l1 = []
                df1 = pd.DataFrame()
                for child in item:
                    l1.append(child.attrib)
                df1 = pd.DataFrame(l1)
                if len(df1.columns) > 0:
                    # Output to Excel for each data object
                    savetoexcel(df1, 'Output.xlsx', 'Table_{}'.format(suffix))


def main():
    # 
    file_val = input("Enter Directory name::")
    
    parser(file_val)
      
      
if __name__ == "__main__":
  
    # calling main function
    main()